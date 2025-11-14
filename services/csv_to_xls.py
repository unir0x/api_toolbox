from flask import send_file
from flask_restx import Namespace, Resource, reqparse, abort
from werkzeug.datastructures import FileStorage
import csv
import io
import os
import re
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from auth.auth import api_auth # Import the new api_auth
import logging

ns = Namespace('csv2xls', description='CSV to XLS operations')

# Define available choices for API parameters
SEPARATOR_CHOICES = ('comma', 'semicolon', 'tab')
TABLE_STYLE_CHOICES = ('TableStyleMedium2', 'TableStyleMedium9', 'TableStyleMedium15', 'TableStyleLight1', 'TableStyleDark1')
LANG_CHOICES = ('SV', 'DA', 'FI', 'NO', 'EN')

parser = reqparse.RequestParser()
parser.add_argument('file', location='files', type=FileStorage, required=True, action='append', help='One or more CSV files to upload')
parser.add_argument('separator', type=str, required=False, default='semicolon', choices=SEPARATOR_CHOICES, help='The separator used in the CSV file.')
parser.add_argument('create_table', type=str, required=False, default='false', choices=('true', 'false'), help='Should the Excel output include a formatted table?')
parser.add_argument('table_style', type=str, required=False, default='TableStyleMedium9', choices=TABLE_STYLE_CHOICES, help='The visual style to apply to the Excel table.')
parser.add_argument('lang', type=str, required=False, default='SV', choices=LANG_CHOICES, help='Language code to determine the default sheet name (e.g., Blad1 for SV).')
parser.add_argument('author', type=str, required=False, default='NorthXL.se', help='The author property to set in the Excel file metadata.')
parser.add_argument('title', type=str, required=False, default='', help='The title property to set in the Excel file metadata.')
parser.add_argument('sheet_name', type=str, required=False, action='append', help='Optional custom sheet names (per file, invalid characters removed).')

# --- Mappings ---
SEPARATOR_MAP = {'comma': ',', 'semicolon': ';', 'tab': '\t'}
SHEET_NAME_MAP = {'SV': 'Blad1', 'DA': 'Ark1', 'FI': 'Taulukko1', 'NO': 'Ark1', 'EN': 'Sheet1'}

@ns.route('')
class CsvToXlsConverter(Resource):
    @ns.expect(parser)
    @ns.doc(security='apiKey')
    @api_auth.login_required
    def post(self):
        """Convert CSV file to Excel with optional table formatting and custom metadata."""
        args = parser.parse_args()
        files = args['file'] or []
        if not isinstance(files, list):
            files = [files]
        files = [f for f in files if f]

        if not files:
            abort(400, 'At least one CSV file must be provided.')

        sep = SEPARATOR_MAP.get(args['separator'], ';')
        base_sheet_template = SHEET_NAME_MAP.get(args['lang'], 'Sheet1')
        sheet_name_inputs = args.get('sheet_name') or []
        if isinstance(sheet_name_inputs, str):
            sheet_name_inputs = [sheet_name_inputs]
        create_table_flag = args['create_table']
        create_table = isinstance(create_table_flag, str) and create_table_flag.lower() == 'true'

        output = io.BytesIO()
        wb = Workbook()
        wb.properties.creator = args['author']
        wb.properties.title = args['title']
        used_sheet_names = set()
        table_counter = 0

        for index, uploaded_file in enumerate(files, start=1):
            if not uploaded_file.filename or not uploaded_file.filename.lower().endswith('.csv'):
                abort(400, f'Invalid file format for "{uploaded_file.filename}". Only .csv files are supported.')

            requested_name = ''
            if isinstance(sheet_name_inputs, list) and len(sheet_name_inputs) >= index:
                requested_name = sheet_name_inputs[index - 1] or ''

            default_sheet_name = default_sheet_name_for_index(base_sheet_template, index)
            sanitized_name = sanitize_sheet_name(requested_name, default_sheet_name)
            sheet_name = ensure_unique_sheet_name(sanitized_name, used_sheet_names)

            ws = wb.active if index == 1 else wb.create_sheet()
            ws.title = sheet_name

            try:
                write_csv_to_sheet(uploaded_file, ws, sep)
            except ValueError as value_error:
                abort(400, str(value_error))
            except Exception as e:
                logging.error(f"Error processing CSV file '{uploaded_file.filename}': {e}")
                abort(400, f"Could not process CSV file '{uploaded_file.filename}'. Please check the file format and the selected separator. Error: {e}")

            if create_table:
                table_counter += 1
                generate_table(ws, args['table_style'], table_counter, sheet_name)
            
        wb.save(output)
        output.seek(0)
        
        if len(files) == 1:
            base_filename = os.path.splitext(files[0].filename)[0] or 'converted_csv'
        else:
            base_name = os.path.splitext(files[0].filename or 'converted')[0]
            base_filename = f"{base_name}_batch"
        new_filename = f"{base_filename}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=new_filename,
            as_attachment=True
        )

def write_csv_to_sheet(file_storage, worksheet, separator):
    """Stream CSV rows into a worksheet and adjust column widths."""
    try:
        file_storage.stream.seek(0)
    except (AttributeError, OSError):
        pass

    decoded_lines = (line.decode('utf-8') for line in file_storage.stream)
    reader = csv.reader(decoded_lines, delimiter=separator)

    column_widths = []
    is_first_row = True

    for row in reader:
        if not row:
            continue

        if is_first_row:
            column_widths = [len(str(cell)) for cell in row]
            is_first_row = False
        else:
            for i, cell in enumerate(row):
                cell_length = len(str(cell))
                if i < len(column_widths):
                    column_widths[i] = max(column_widths[i], cell_length)
                else:
                    column_widths.append(cell_length)

        worksheet.append(row)

    if is_first_row:
        raise ValueError(f'The provided CSV file "{file_storage.filename}" is empty.')

    adjust_column_widths(worksheet, column_widths)


def adjust_column_widths(ws, column_widths):
    """Adjust column widths based on the longest value found in each column."""
    extra_space = 4
    for i, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = width + extra_space

def generate_table(ws, table_style_name, index, sheet_title):
    """Create a table in the worksheet with the specified style."""
    # Check if worksheet is empty before creating a table
    if ws.max_row == 0:
        return
    table_name = build_table_name(sheet_title, index)
    tab = Table(displayName=table_name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    style = TableStyleInfo(
        name=table_style_name, 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

INVALID_SHEET_CHARS = re.compile(r'[:\\/?*\[\]]')
TABLE_NAME_INVALID_CHARS = re.compile(r'[^A-Za-z0-9_]')  # Excel table names allow letters, numbers, underscore

def sanitize_sheet_name(name, fallback):
    """Return a sheet name that obeys Excel constraints, otherwise fallback."""
    if not name:
        return fallback
    cleaned = INVALID_SHEET_CHARS.sub('', name).strip()
    if not cleaned:
        return fallback
    return cleaned[:31]

def ensure_unique_sheet_name(name, used_names):
    """Ensure the sheet name is unique within the workbook."""
    candidate = name
    suffix = 1
    while candidate in used_names:
        trimmed = name[: max(0, 31 - len(str(suffix)) - 1)]
        candidate = f"{trimmed}_{suffix}" if trimmed else f"{name}_{suffix}"
        candidate = candidate[:31]
        suffix += 1
    used_names.add(candidate)
    return candidate

def default_sheet_name_for_index(base_name, index):
    """Derive a sheet name from the language default that increments per file."""
    match = re.match(r'^(.*?)(\d+)$', base_name)
    if match:
        prefix, start_number = match.groups()
        start = int(start_number)
        return f"{prefix}{start + index - 1}"
    return f"{base_name}{index}"

def build_table_name(sheet_title, index):
    """Generate a workbook-unique table name based on sheet title."""
    base = TABLE_NAME_INVALID_CHARS.sub('', sheet_title) or 'DataTable'
    name = f"{base}_{index}"
    # Excel table names must start with a letter; prefix if needed
    if not name[0].isalpha():
        name = f"T{name}"
    return name[:31]
